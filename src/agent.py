import logging
import sys
import os
from dotenv import load_dotenv
from system_prompt import SYSTEM_PROMPT
import json
import openpyxl
from rag import pinecone_search as _rag_pinecone_search

load_dotenv()

# Ensure logs are visible
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
    stream=sys.stdout,
    force=True,
)

from decouple import config
from livekit.agents import (
    JobContext,
    AgentServer,
    cli,
    llm,
    room_io,
)
from livekit.agents.voice import Agent, AgentSession
from livekit.plugins import openai, deepgram, elevenlabs, silero

try:
    from livekit.plugins import noise_cancellation
    _HAS_NOISE_CANCELLATION = True
except ImportError:
    _HAS_NOISE_CANCELLATION = False


logger = logging.getLogger("voice-agent")

# Required env vars for STT, TTS, LLM, RAG - fail fast if missing
_REQUIRED_KEYS = [
    ("DEEPGRAM_API_KEY", "STT (speech-to-text)"),
    ("ELEVEN_API_KEY", "TTS (text-to-speech)"),
    ("GROQ_API_KEY", "LLM"),
    ("PINECONE_API_KEY", "RAG search"),
    ("PINECONE_HOST", "RAG search"),
    ("PINECONE_NAMESPACE", "RAG search"),
]


def _validate_env():
    """Validate required API keys at startup. Exits with clear error if any missing."""
    missing = []
    for key, desc in _REQUIRED_KEYS:
        val = config(key, default="")
        if not val or not str(val).strip():
            missing.append(f"  {key} ({desc})")
    if missing:
        msg = "Missing required env vars. Set them in .env:\n" + "\n".join(missing)
        print(f"[PHARMA] >>> ERROR: {msg}", flush=True)
        raise SystemExit(1)


server = AgentServer()


def prewarm(proc):
    """Prewarm VAD for faster startup."""
    proc.userdata["vad"] = silero.VAD.load()


server.setup_fnc = prewarm

# Load the database once
DB_PATH = "data/db.json"
try:
    with open(DB_PATH, "r") as f:
        PATIENT_DB = json.load(f)
    print(f"[PHARMA] >>> Loaded {len(PATIENT_DB)} records from {DB_PATH}", flush=True)
except Exception as e:
    print(f"[PHARMA] >>> ERROR loading DB: {e}", flush=True)
    PATIENT_DB = []

# Load the drug code Excel file once
DRUG_CODE_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "data", "Claim Drug Code List.xlsx")
DRUG_CODE_DB = []
try:
    wb = openpyxl.load_workbook(DRUG_CODE_PATH, read_only=True, data_only=True)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        record = dict(zip(headers, row))
        DRUG_CODE_DB.append(record)
    wb.close()
    print(f"[PHARMA] >>> Loaded {len(DRUG_CODE_DB)} drug codes from {DRUG_CODE_PATH}", flush=True)
except Exception as e:
    print(f"[PHARMA] >>> ERROR loading drug code list: {e}", flush=True)
    DRUG_CODE_DB = []


def _search_drug_codes(drug_code: str | None = None, drug_name: str | None = None, max_results: int = 5) -> list[dict]:
    """Search the in-memory drug code database by code or name (brand/scientific)."""
    matches = []
    for record in DRUG_CODE_DB:
        # Exact match on drug code
        if drug_code:
            code_val = str(record.get("Code", "")).strip()
            if drug_code.strip().lower() == code_val.lower():
                matches.append(record)
                continue

        # Partial case-insensitive match on name fields
        if drug_name:
            search_term = drug_name.strip().lower()
            scientific = str(record.get("Scientific Name", "")).lower()
            brand = str(record.get("Description", "")).lower()
            if search_term in scientific or search_term in brand:
                matches.append(record)
                continue

        if len(matches) >= max_results:
            break

    # Format results for the agent
    results = []
    for m in matches[:max_results]:
        results.append({
            "drug_code": m.get("Code", ""),
            "scientific_name": m.get("Scientific Name", ""),
            "brand_name": m.get("Description", ""),
            "strength": m.get("Strength", ""),
            "route": m.get("Roa", ""),
            "dosage_form": m.get("Dosage Form Package", ""),
            "unit_price_aed": m.get("Price", ""),
            "package_size": m.get("Package Size", ""),
            "active": m.get("Active", ""),
        })
    return results

@server.rtc_session(agent_name="pharmacy-agent")
async def pharmacy_agent(ctx: JobContext):
    print(f"[PHARMA] >>> Entrypoint called for room {ctx.room.name}", flush=True)
    logger.info(f"Starting agent for room {ctx.room.name}")

    try:
        initial_ctx = llm.ChatContext()
        
        # Use shared system prompt from rag.py
        initial_ctx.add_message(
            content=SYSTEM_PROMPT,
            role="system",
        )

        class PharmacyTools:
            @llm.function_tool(
                description="Semantic search over the insurance and pharmacy database. Use this to find similar past cases, check general policy rules, or search when you don't have an exact identifier. Returns patient records including policy details, medication coverage, claim status, denial codes, dispensing history, and alternative drug availability."
            )
            async def pinecone_search(self, query: str, top_k: int = 3):
                logger.info(f"Searching Pinecone for: {query}")
                try:
                    results = _rag_pinecone_search(query, top_k)
                except Exception as e:
                    logger.exception("RAG pinecone_search failed")
                    print(f"[PHARMA] >>> RAG ERROR: {e}", flush=True)
                    return f"Search failed: {e}. Please try rephrasing or ask without database lookup."
                if not results:
                    return "No relevant records found."
                return "\n\n".join(results)

            @llm.function_tool(
                description="Look up a specific patient record by Emirates ID, Policy Number, Member Card Number, Claim ID, Patient ID, or Patient Name. Use this FIRST when you have a specific identifier. Returns: patient identity, insurance policy details (plan, copay, limits, active status), prescription details, dispensing history (prior dispenses, already dispensed this cycle), claim status, denial code and reason, recommended resolution, inventory status, and alternative drugs with availability."
            )
            async def lookup_database(
                self,
                emirates_id: str | None = None,
                policy_number: str | None = None,
                member_card_number: str | None = None,
                claim_id: str | None = None,
                patient_id: str | None = None,
                patient_name: str | None = None,
            ):
                """
                Retrieves a patient record from the local database by exact match on identifiers.
                member_card_number is treated as a policy number lookup.
                For patient_name, performs a case-insensitive partial match.
                """
                # member_card_number maps to policy_number in the DB
                effective_policy = policy_number or member_card_number

                logger.info(
                    f"DB Lookup: eid={emirates_id}, pol={effective_policy}, clm={claim_id}, pid={patient_id}, name={patient_name}"
                )
                
                matches = []
                for record in PATIENT_DB:
                    # check EMIRATES ID
                    if emirates_id and record.get("emirates_id") == emirates_id:
                        matches.append(record)
                        continue
                    
                    # check POLICY NUMBER / MEMBER CARD NUMBER
                    if effective_policy and record.get("policy_number") == effective_policy:
                        matches.append(record)
                        continue

                    # check CLAIM ID
                    if claim_id and record.get("claim_id") == claim_id:
                        matches.append(record)
                        continue

                    # check PATIENT ID
                    if patient_id and record.get("patient_id") == patient_id:
                        matches.append(record)
                        continue

                    # check NAME (partial match)
                    if patient_name:
                        rec_name = record.get("patient_name", "").lower()
                        if patient_name.lower() in rec_name:
                            matches.append(record)
                            continue

                if not matches:
                    return "No records found matching the provided details."

                # Return JSON string of matches (limit to top 3 to avoid context overflow)
                return json.dumps(matches[:3], indent=2)

            @llm.function_tool(
                description="Look up a drug by its drug code (e.g. '0005-116801-1161') or by drug name (brand or scientific/generic name). Returns the official drug code, scientific name, brand name, strength, route, dosage form, unit price in AED, and active/discontinued status. Use this when a caller mentions a medication by name and you need to verify its drug code, price, or availability."
            )
            async def lookup_drug_code(
                self,
                drug_code: str | None = None,
                drug_name: str | None = None,
            ):
                """
                Searches the drug code database.
                Provide either drug_code for exact code lookup, or drug_name for partial name search.
                """
                logger.info(f"Drug code lookup: code={drug_code}, name={drug_name}")

                if not drug_code and not drug_name:
                    return "Please provide either a drug code or a drug name to search."

                results = _search_drug_codes(drug_code=drug_code, drug_name=drug_name)

                if not results:
                    return "No matching drugs found in the drug code database. Please verify the drug code or name."

                return json.dumps(results, indent=2)

        pharmacy_tools = PharmacyTools()
        vad = ctx.proc.userdata.get("vad") or silero.VAD.load()

        session = AgentSession(
            stt=deepgram.STT(
                model="nova-3",
                language="en-US",
                api_key=config("DEEPGRAM_API_KEY"),
            ),
            llm=openai.LLM(
                base_url="https://api.groq.com/openai/v1",
                api_key=config("GROQ_API_KEY"),
                model="openai/gpt-oss-120b",
            ),
            tts=elevenlabs.TTS(
                api_key=config("ELEVEN_API_KEY"),
                voice_id="i80JxxvpWr5Q7cTdT1Ik",
                # voice_settings=elevenlabs.VoiceSettings(
                #     stability=0.5,
                #     similarity_boost=0.75,
                #     speed=0.8  # Adjust this between 0.7 and 1.2
                # )
            ),
            vad=vad,
            turn_detection="vad",
            allow_interruptions=True,
            tools=[pharmacy_tools.pinecone_search, pharmacy_tools.lookup_database, pharmacy_tools.lookup_drug_code],
        )

        class PharmacyAgent(Agent):
            async def on_enter(self) -> None:
                """Greet immediately via direct TTS - publishes audio track for playground."""
                print("[PHARMA] >>> on_enter called, saying greeting", flush=True)
                self.session.say("Hello! This is a pharmacy insurance approval agent. Please provide the patient's name or ID and the drug class you're inquiring about.")

        agent = PharmacyAgent(
            instructions="You are a professional pharmacy insurance approval agent. Verify drug coverage based on patient tiers.",
            chat_ctx=initial_ctx,
        )

        room_opts = room_io.RoomOptions()
        if _HAS_NOISE_CANCELLATION:
            room_opts = room_io.RoomOptions(
                audio_input=room_io.AudioInputOptions(
                    noise_cancellation=noise_cancellation.BVC(),
                ),
            )

        print("[PHARMA] >>> Starting session (connects to room automatically)...", flush=True)
        await session.start(
            agent=agent,
            room=ctx.room,
            room_options=room_opts,
        )
        print(f"[PHARMA] >>> Agent session started for room {ctx.room.name}", flush=True)
        logger.info(f"Agent session started for room {ctx.room.name}")
    except Exception as e:
        print(f"[PHARMA] >>> ERROR: {e}", flush=True)
        logger.exception("Agent failed")
        raise


if __name__ == "__main__":
    if len(sys.argv) > 1 and sys.argv[1] == "download-files":
        print("Downloading Silero VAD model...", flush=True)
        silero.VAD.load()
        print("Done.", flush=True)
        sys.exit(0)
    _validate_env()
    print("[PHARMA] >>> Env validation OK (STT, TTS, LLM, RAG keys present)", flush=True)
    cli.run_app(server)